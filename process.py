"""
process.py — Filter and validate CSD minimal pairs dataset files.

Subcommands
-----------
  python process.py filter --excel data/review/reviews_dataset.xlsx
  python process.py validate
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

ROOT          = Path(__file__).parent
DATA_DIR      = ROOT / "data"
GENERATED_DIR = DATA_DIR / "generated"
FILTERED_DIR  = DATA_DIR / "filtered"

VARIANT_ORDER = {"A": 0, "B": 1, "C": 2, "D1": 3, "D2": 4, "D3": 5}

# Exact mapping from Excel sheet name to source JSONL filename
SHEET_TO_JSONL = {
    "type1_2np": "type1_2np.jsonl",
    "type1_3np": "type1_3np.jsonl",
    "type1_4np": "type1_4np.jsonl",
    "type2_2np": "type2_2np.jsonl",
    "type2_3np": "type2_3np.jsonl",
    "type3_2np": "type3_2np.jsonl",
}

MINIMUM_TARGETS = {
    "type1_2np": {"perception": 25, "causative": 14},
    "type2_2np": {"perception": 25, "causative": 14},
    "type3_2np": {"perception": 25, "causative": 14},
    "type1_3np": {"perception": 15, "causative": 8},
    "type2_3np": {"perception": 15, "causative": 8},
    "type1_4np": {"perception": 15},
}

"""
HELPERS
"""
def _base_item(pair_id: str) -> str:
    return pair_id.rsplit("_", 1)[0]


def _variant(pair_id: str) -> str:
    return pair_id.rsplit("_", 1)[1]


def _sort_key(pair_id: str):
    """Sort key: base item prefix, then item number, then variant order."""
    base   = _base_item(pair_id)
    var    = _variant(pair_id)
    parts  = base.split("_")
    num    = int(parts[-1])
    prefix = "_".join(parts[:-1])
    return (prefix, num, VARIANT_ORDER.get(var, 99))


def _renumber(records: list[dict]) -> list[dict]:
    """Reassign pair_id item numbers sequentially (001..N) with no gaps.

    Preserves current record order to determine assignment; sort before calling
    so that numbers reflect the intended final order.
    """
    seen: dict[str, int] = {}
    counter = 0
    for rec in records:
        base = _base_item(rec["pair_id"])
        if base not in seen:
            counter += 1
            seen[base] = counter

    out = []
    for rec in records:
        base, var = rec["pair_id"].rsplit("_", 1)
        prefix    = "_".join(base.split("_")[:-1])
        new_base  = f"{prefix}_{seen[base]:03d}"
        new_rec   = dict(rec)
        new_rec["pair_id"] = f"{new_base}_{var}"
        out.append(new_rec)
    return out


def _load_jsonl(path: Path) -> list[dict]:
    records = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    return records


def _write_jsonl(records: list[dict], path: Path) -> None:
    with open(path, "w", encoding="utf-8") as f:
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def _read_sheet_annotations(ws) -> tuple[dict[str, str], dict[str, str]]:
    """Read a single openpyxl worksheet and return two dicts keyed by base_item_id.

    Returns:
        exclude_bases  — {base_id: pair_id}  for rows where exclude == "exclude"
        note_bases     — {base_id: exclude_reason} for rows where exclude == "note"

    Only Variant A rows are ever annotated; non-A rows and blank exclude cells
    are silently skipped.
    """
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        return {}, {}

    # Build column index from header; handle None cells and strip whitespace
    col_idx: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        if cell is not None:
            col_idx[str(cell).strip()] = i

    def get(row, col):
        idx = col_idx.get(col)
        if idx is None or idx >= len(row):
            return ""
        val = row[idx]
        return str(val).strip() if val is not None else ""

    exclude_bases: dict[str, str] = {}
    note_bases:    dict[str, str] = {}

    for row in rows_iter:
        pair_id = get(row, "pair_id")
        v_type  = get(row, "v_type")
        exclude = get(row, "exclude").lower()
        reason  = get(row, "exclude_reason")

        if v_type != "A" or not exclude:
            continue

        base = _base_item(pair_id)

        if exclude == "exclude":
            exclude_bases[base] = {
                "pair_id": pair_id,
                "reason": reason
            }
        elif exclude == "note":
            note_bases[base] = reason

    return exclude_bases, note_bases


"""
FILTER
"""
def cmd_filter(args):
    if openpyxl is None:
        sys.exit(
            "Error: openpyxl is not installed.\n"
            "Install it with:  pip install openpyxl"
        )

    excel_path = Path(args.excel)
    if not excel_path.is_absolute():
        excel_path = ROOT / excel_path
    if not excel_path.exists():
        sys.exit(f"Error: Excel file not found: {excel_path}")

    FILTERED_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    all_excluded_records: list[dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_TO_JSONL:
            # Non-condition sheets (e.g. Sheet1) are skipped silently
            continue

        jsonl_name = SHEET_TO_JSONL[sheet_name]
        jsonl_path = GENERATED_DIR / jsonl_name
        if not jsonl_path.exists():
            print(f"Warning: {jsonl_name} not found in data/generated/ — skipping sheet '{sheet_name}'")
            continue

        ws = wb[sheet_name]
        exclude_bases, note_bases = _read_sheet_annotations(ws)

        records = _load_jsonl(jsonl_path)

        # Build index: base_id → list of records (all variants)
        base_to_records: dict[str, list[dict]] = {}
        for rec in records:
            base = _base_item(rec["pair_id"])
            base_to_records.setdefault(base, []).append(rec)

        # Apply exclusions — drop entire base items
        excluded_bases = set(exclude_bases.keys()) & set(base_to_records.keys())
        remaining: list[dict] = []
        for rec in records:
            base = _base_item(rec["pair_id"])
            if base in excluded_bases:
                # Collect for exclusion log with provenance fields
                log_rec = dict(rec)
                log_rec["excluded_from"]    = sheet_name
                log_rec["exclusion_reason"] = exclude_bases[base]["reason"] or ""
                all_excluded_records.append(log_rec)
            else:
                remaining.append(rec)

        # Apply note patches — overwrite notes field for every variant
        note_patches_applied = 0
        for rec in remaining:
            base = _base_item(rec["pair_id"])
            if base in note_bases:
                manual_note = f"Manual revision note: {note_bases[base]}"
                
                existing_note = rec.get("notes")
                if existing_note:
                    rec["notes"] = f"{existing_note} | {manual_note}"
                else:
                    rec["notes"] = manual_note

                if _variant(rec["pair_id"]) == "A":
                    note_patches_applied += 1

        # Renumber and sort
        remaining.sort(key=lambda r: _sort_key(r["pair_id"]))
        remaining = _renumber(remaining)
        remaining.sort(key=lambda r: _sort_key(r["pair_id"]))

        # Summary counts
        remaining_bases: set[str] = {_base_item(r["pair_id"]) for r in remaining}
        a_by_class: dict[str, int] = {}
        for r in remaining:
            if _variant(r["pair_id"]) == "A":
                cls = r.get("V1_class", "unknown")
                a_by_class[cls] = a_by_class.get(cls, 0) + 1

        out_path = FILTERED_DIR / f"{sheet_name}_filtered.jsonl"
        _write_jsonl(remaining, out_path)

        class_str = ", ".join(f"{k}={v}" for k, v in sorted(a_by_class.items()))
        print(
            f"{sheet_name}: excluded {len(excluded_bases)} base items, "
            f"{note_patches_applied} note patches applied, "
            f"{len(remaining_bases)} items remain | "
            f"Variant A by class: {class_str}"
        )
        print(f"  Written {len(remaining)} records → {out_path}")

    wb.close()

    # Write exclusion log
    all_excluded_records.sort(
        key=lambda r: (r.get("excluded_from", ""), _sort_key(r["pair_id"]))
    )
    excl_path = FILTERED_DIR / "excluded_items.jsonl"
    _write_jsonl(all_excluded_records, excl_path)
    print(f"\nExclusion log: {len(all_excluded_records)} records → {excl_path}")



"""
VALIDATE DATASET
"""
def cmd_validate(_args):
    if not FILTERED_DIR.exists():
        sys.exit(
            "Error: data/filtered/ does not exist. "
            "Run 'python process.py filter --excel <path>' first."
        )

    filtered_files = sorted(FILTERED_DIR.glob("*_filtered.jsonl"))
    if not filtered_files:
        sys.exit(
            "Error: No *_filtered.jsonl files found in data/filtered/. "
            "Run 'python process.py filter --excel <path>' first."
        )

    all_pass = True
    for path in filtered_files:
        records  = _load_jsonl(path)
        failures: list[str] = []

        a_records = [r for r in records if _variant(r["pair_id"]) == "A"]

        # 1. Variant A counts by class and lemma
        a_by_class: dict[str, int] = {}
        a_by_lemma: dict[str, int] = {}
        for r in a_records:
            cls   = r.get("V1_class", "unknown")
            lemma = r.get("V1_lemma", "unknown")
            a_by_class[cls]   = a_by_class.get(cls, 0) + 1
            a_by_lemma[lemma] = a_by_lemma.get(lemma, 0) + 1

        # 2. Sequential item numbers — print min, max, total; flag gaps
        seen_bases: list[str] = []
        seen_set: set[str] = set()
        for r in sorted(records, key=lambda r: _sort_key(r["pair_id"])):
            b = _base_item(r["pair_id"])
            if b not in seen_set:
                seen_set.add(b)
                seen_bases.append(b)
        nums     = sorted(int(b.split("_")[-1]) for b in seen_bases)
        min_num  = nums[0]  if nums else 0
        max_num  = nums[-1] if nums else 0
        total    = len(nums)
        expected = list(range(min_num, max_num + 1))
        if nums != expected:
            failures.append(
                f"Item numbers not sequential: min={min_num}, max={max_num}, "
                f"total={total}, expected {max_num - min_num + 1} contiguous values"
            )

        # 3. Duplicate grammatical sentences among Variant A
        gram_seen: set[str] = set()
        dup_grams: list[str] = []
        for r in a_records:
            s = r.get("grammatical", "")
            if s in gram_seen:
                dup_grams.append(s)
            gram_seen.add(s)
        if dup_grams:
            failures.append(
                f"Duplicate grammatical sentences ({len(dup_grams)}): "
                + "; ".join(f'"{s}"' for s in dup_grams[:3])
            )

        # 4. Duplicate pair_ids
        id_counts: dict[str, int] = {}
        for r in records:
            pid = r["pair_id"]
            id_counts[pid] = id_counts.get(pid, 0) + 1
        dup_ids = [pid for pid, cnt in id_counts.items() if cnt > 1]
        if dup_ids:
            failures.append(
                f"Duplicate pair_ids ({len(dup_ids)}): {dup_ids[:5]}"
            )

        # 5. Minimum Variant A targets
        # Strip the _filtered suffix to recover the condition key
        condition = path.stem.removesuffix("_filtered")
        if condition in MINIMUM_TARGETS:
            for cls, min_count in MINIMUM_TARGETS[condition].items():
                actual = a_by_class.get(cls, 0)
                if actual < min_count:
                    failures.append(
                        f"Below minimum for '{cls}': {actual} < {min_count}"
                    )

        verdict   = "PASS" if not failures else "FAIL"
        class_str = ", ".join(f"{k}={v}" for k, v in sorted(a_by_class.items()))
        lemma_str = ", ".join(f"{k}={v}" for k, v in sorted(a_by_lemma.items()))
        print(
            f"[{verdict}] {path.name} — {len(records)} records | "
            f"Variant A: {class_str} | lemmas: {lemma_str}"
        )
        print(
            f"         item numbers: min={min_num}, max={max_num}, "
            f"total unique={total}"
        )
        for msg in failures:
            print(f"         FAIL: {msg}")

        if failures:
            all_pass = False

    print()
    print("All files PASS." if all_pass else "One or more files FAIL — review the issues above.")


"""
CLI ENTRY POINT
"""
def main():
    parser = argparse.ArgumentParser(
        description="Post-processing pipeline for CSD minimal pairs dataset.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p_filter = sub.add_parser(
        "filter",
        help="Apply Excel review annotations, renumber, write to data/filtered/.",
    )
    p_filter.add_argument(
        "--excel",
        required=True,
        metavar="PATH",
        help="Path to the review Excel file (e.g. data/review/reviews_dataset.xlsx).",
    )

    sub.add_parser(
        "validate",
        help="Validate *_filtered.jsonl files in data/filtered/ .",
    )

    args = parser.parse_args()

    if args.command == "filter":
        cmd_filter(args)
    elif args.command == "validate":
        cmd_validate(args)


if __name__ == "__main__":
    main()
